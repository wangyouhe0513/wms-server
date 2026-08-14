"""
工厂进销存管理系统 — API 后端
"""
import hashlib
import secrets
import io
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import List, Optional

import openpyxl
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Query, Header
from fastapi.responses import StreamingResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, extract, or_

from database import init_db, get_db
from models import ProductSpec, Color, ProductSku, Salesperson, Transaction, InventoryLog, Admin, OperationLog, SystemConfig, FinanceRecord, FinanceCategory, SalaryWorker, SalaryPrice, SalaryRecord, SalarySpecItem, SalaryViewLog

app = FastAPI(title="工厂进销存管理系统")

# Session token store: {token: (admin_id, created_at)}
_active_tokens = {}
TOKEN_EXPIRE_HOURS = 24


# ============================================================
# 工具函数
# ============================================================
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def generate_token() -> str:
    return secrets.token_hex(32)


def log_operation(db: Session, admin_id: int, admin_name: str, action: str, target: str, detail: str = "", ip: str = ""):
    log = OperationLog(
        admin_id=admin_id, admin_name=admin_name,
        action=action, target=target, detail=detail, ip_address=ip
    )
    db.add(log)


def get_config(db: Session, key: str, default: str = "") -> str:
    """读取系统配置值"""
    config = db.query(SystemConfig).filter(SystemConfig.key == key).first()
    return config.value if config else default


def set_config(db: Session, key: str, value: str):
    """写入系统配置值"""
    config = db.query(SystemConfig).filter(SystemConfig.key == key).first()
    if config:
        config.value = value
    else:
        config = SystemConfig(key=key, value=value)
        db.add(config)


def get_low_stock_threshold(db: Session) -> int:
    """获取全局默认低库存阈值"""
    val = get_config(db, "low_stock_threshold", "50")
    try:
        return int(val)
    except (ValueError, TypeError):
        return 50


def get_current_admin(authorization: str = Header(None), db: Session = Depends(get_db)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "请先登录")
    token = authorization[7:]

    # 先从内存查找（兼容旧方式）
    entry = _active_tokens.get(token)
    if entry:
        admin_id, created_at = entry
        if (datetime.now() - created_at).total_seconds() > TOKEN_EXPIRE_HOURS * 3600:
            del _active_tokens[token]
            raise HTTPException(401, "登录已过期（24小时），请重新登录")
        admin = db.query(Admin).filter(Admin.id == admin_id, Admin.is_active == 1).first()
        if admin: return admin
        raise HTTPException(401, "账户已被禁用")

    # DB持久化token（服务器重启不丢）
    admin = db.query(Admin).filter(Admin.token == token, Admin.is_active == 1).first()
    if not admin:
        raise HTTPException(401, "请先登录")
    if admin.token_expires_at and datetime.now() > admin.token_expires_at:
        admin.token = None; admin.token_expires_at = None
        db.commit()
        raise HTTPException(401, "登录已过期（24小时），请重新登录")
    return admin


# ============================================================
# 启动时初始化数据库
# ============================================================
@app.on_event("startup")
def startup():
    init_db()
    db = next(get_db())
    try:
        # 创建默认管理员（首次启动）
        existing = db.query(Admin).filter(Admin.username == "admin").first()
        if not existing:
            admin = Admin(username="admin", password_hash=hash_password("admin123"), role="superadmin")
            db.add(admin)
            db.commit()
            print("✅ 默认管理员已创建: admin / admin123")
    finally:
        db.close()


# ============================================================
# 认证 API — 登录 / 登出
# ============================================================
from pydantic import BaseModel as PydanticBaseModel


class LoginRequest(PydanticBaseModel):
    username: str
    password: str


@app.post("/api/auth/login")
def login(req: LoginRequest, db: Session = Depends(get_db)):
    admin = db.query(Admin).filter(Admin.username == req.username, Admin.is_active == 1).first()
    if not admin or admin.password_hash != hash_password(req.password):
        raise HTTPException(401, "用户名或密码错误")
    token = generate_token()
    _active_tokens[token] = (admin.id, datetime.now())
    # 持久化到DB，服务器重启不丢登录状态
    admin.token = token
    admin.token_expires_at = datetime.now() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    admin.last_login = datetime.now()
    log_operation(db, admin.id, admin.username, "login", "auth", f"管理员登录")
    db.commit()
    return {
        "token": token,
        "username": admin.username,
        "role": admin.role,
    }


@app.post("/api/auth/logout")
def logout(admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    # 清除内存+DB token
    expired = [k for k, v in _active_tokens.items() if v[0] == admin.id]
    for k in expired:
        del _active_tokens[k]
    admin.token = None
    admin.token_expires_at = None
    log_operation(db, admin.id, admin.username, "logout", "auth", "管理员登出")
    db.commit()
    return {"ok": True}


@app.get("/api/auth/me")
def me(admin: Admin = Depends(get_current_admin)):
    return {"username": admin.username, "role": admin.role}


# ============================================================
# 管理员 CRUD API（仅 superadmin 可操作）
# ============================================================
@app.get("/api/admins")
def list_admins(admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if admin.role != "superadmin":
        raise HTTPException(403, "仅超级管理员可操作")
    admins = db.query(Admin).filter(Admin.is_active == 1).all()
    return [{"id": a.id, "username": a.username, "role": a.role, "last_login": str(a.last_login) if a.last_login else None, "created_at": str(a.created_at)} for a in admins]


@app.post("/api/admins")
def create_admin(req: LoginRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if admin.role != "superadmin":
        raise HTTPException(403, "仅超级管理员可操作")
    existing = db.query(Admin).filter(Admin.username == req.username).first()
    if existing:
        if existing.is_active == 0:
            existing.is_active = 1
            existing.password_hash = hash_password(req.password)
            db.commit()
            log_operation(db, admin.id, admin.username, "create", "admin", f"重新激活管理员: {req.username}")
            return {"ok": True, "reactivated": True}
        raise HTTPException(400, "用户名已存在")
    new_admin = Admin(username=req.username, password_hash=hash_password(req.password), role="admin")
    db.add(new_admin)
    db.commit()
    log_operation(db, admin.id, admin.username, "create", "admin", f"创建管理员: {req.username}")
    return {"ok": True, "id": new_admin.id}


class ResetPasswordRequest(PydanticBaseModel):
    password: str


@app.put("/api/admins/{admin_id}/password")
def reset_password(admin_id: int, req: ResetPasswordRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if admin.role != "superadmin":
        raise HTTPException(403, "仅超级管理员可操作")
    target = db.query(Admin).filter(Admin.id == admin_id).first()
    if not target:
        raise HTTPException(404, "管理员不存在")
    target.password_hash = hash_password(req.password)
    db.commit()
    log_operation(db, admin.id, admin.username, "update", "admin", f"重置密码: {target.username}")
    return {"ok": True}


@app.delete("/api/admins/{admin_id}")
def delete_admin(admin_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if admin.role != "superadmin":
        raise HTTPException(403, "仅超级管理员可操作")
    if admin_id == admin.id:
        raise HTTPException(400, "不能删除自己")
    target = db.query(Admin).filter(Admin.id == admin_id).first()
    if not target:
        raise HTTPException(404, "管理员不存在")
    target.is_active = 0
    db.commit()
    # Remove their tokens
    expired = [k for k, v in _active_tokens.items() if v[0] == admin_id]
    for k in expired:
        del _active_tokens[k]
    log_operation(db, admin.id, admin.username, "delete", "admin", f"删除管理员: {target.username}")
    return {"ok": True}


@app.put("/api/auth/change-password")
def change_password(old_password: str, new_password: str, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if admin.password_hash != hash_password(old_password):
        raise HTTPException(400, "原密码错误")
    admin.password_hash = hash_password(new_password)
    db.commit()
    log_operation(db, admin.id, admin.username, "update", "auth", "修改密码")
    return {"ok": True}


# ============================================================
# 操作日志 API
# ============================================================
@app.get("/api/logs")
def get_logs(limit: int = 100, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    logs = db.query(OperationLog).order_by(OperationLog.created_at.desc()).limit(limit).all()
    return [{
        "id": l.id,
        "admin_name": l.admin_name,
        "action": l.action,
        "target": l.target,
        "detail": l.detail,
        "ip_address": l.ip_address,
        "created_at": str(l.created_at),
    } for l in logs]
@app.get("/api/specs")
def list_specs(db: Session = Depends(get_db)):
    specs = db.query(ProductSpec).filter(ProductSpec.is_active == 1).order_by(ProductSpec.sort_order, ProductSpec.id).all()
    return [{"id": s.id, "name": s.name, "sort_order": s.sort_order} for s in specs]


@app.post("/api/specs")
def create_spec(name: str, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    existing = db.query(ProductSpec).filter(ProductSpec.name == name).first()
    if existing:
        if existing.is_active == 0:
            existing.is_active = 1
            log_operation(db, admin.id, admin.username, "create", "spec", f"重新激活规格: {name}")
            db.commit()
            return {"id": existing.id, "name": existing.name, "reactivated": True}
        raise HTTPException(400, "规格已存在")
    spec = ProductSpec(name=name)
    db.add(spec)
    log_operation(db, admin.id, admin.username, "create", "spec", f"创建规格: {name}")
    db.commit()
    db.refresh(spec)
    return {"id": spec.id, "name": spec.name}


@app.put("/api/specs/{spec_id}")
def update_spec(spec_id: int, name: str, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    spec = db.query(ProductSpec).filter(ProductSpec.id == spec_id).first()
    if not spec:
        raise HTTPException(404, "规格不存在")
    old = spec.name
    spec.name = name
    log_operation(db, admin.id, admin.username, "update", "spec", f"修改规格: {old} -> {name}")
    db.commit()
    return {"ok": True}


@app.delete("/api/specs/{spec_id}")
def delete_spec(spec_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    spec = db.query(ProductSpec).filter(ProductSpec.id == spec_id).first()
    if not spec:
        raise HTTPException(404, "规格不存在")

    # 检查是否有库存或交易记录的 SKU
    has_stock = db.query(ProductSku).filter(
        ProductSku.spec_id == spec_id,
        ProductSku.current_stock > 0
    ).first()
    if has_stock:
        raise HTTPException(400, f"规格「{spec.name}」下还有库存，不能删除")

    has_txn = db.query(Transaction).join(ProductSku).filter(
        ProductSku.spec_id == spec_id
    ).first()
    if has_txn:
        raise HTTPException(400, f"规格「{spec.name}」下有交易记录，不能删除")

    spec.is_active = 0
    log_operation(db, admin.id, admin.username, "delete", "spec", f"删除规格: {spec.name}")
    db.commit()
    return {"ok": True}


# ============================================================
# 基础数据管理 — 颜色
# ============================================================
@app.get("/api/colors")
def list_colors(db: Session = Depends(get_db)):
    colors = db.query(Color).filter(Color.is_active == 1).order_by(Color.sort_order, Color.id).all()
    return [{"id": c.id, "name": c.name, "code": c.code} for c in colors]


@app.post("/api/colors")
def create_color(name: str, code: str = "", admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    existing = db.query(Color).filter(Color.name == name).first()
    if existing:
        if existing.is_active == 0:
            existing.is_active = 1
            log_operation(db, admin.id, admin.username, "create", "color", f"重新激活颜色: {name}")
            db.commit()
            return {"id": existing.id, "name": existing.name, "reactivated": True}
        raise HTTPException(400, "颜色已存在")
    color = Color(name=name, code=code)
    db.add(color)
    log_operation(db, admin.id, admin.username, "create", "color", f"创建颜色: {name}")
    db.commit()
    db.refresh(color)
    return {"id": color.id, "name": color.name}


@app.put("/api/colors/{color_id}")
def update_color(color_id: int, name: str, code: str = "", admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    color = db.query(Color).filter(Color.id == color_id).first()
    if not color:
        raise HTTPException(404, "颜色不存在")
    old = color.name
    color.name = name
    color.code = code
    log_operation(db, admin.id, admin.username, "update", "color", f"修改颜色: {old} -> {name}")
    db.commit()
    return {"ok": True}


@app.delete("/api/colors/{color_id}")
def delete_color(color_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    color = db.query(Color).filter(Color.id == color_id).first()
    if not color:
        raise HTTPException(404, "颜色不存在")

    # 检查是否有库存或交易记录的 SKU
    has_stock = db.query(ProductSku).filter(
        ProductSku.color_id == color_id,
        ProductSku.current_stock > 0
    ).first()
    if has_stock:
        raise HTTPException(400, f"颜色「{color.name}」下还有库存，不能删除")

    has_txn = db.query(Transaction).join(ProductSku).filter(
        ProductSku.color_id == color_id
    ).first()
    if has_txn:
        raise HTTPException(400, f"颜色「{color.name}」下有交易记录，不能删除")

    color.is_active = 0
    log_operation(db, admin.id, admin.username, "delete", "color", f"删除颜色: {color.name}")
    db.commit()
    return {"ok": True}


# ============================================================
# 基础数据管理 — 销售人员
# ============================================================
@app.get("/api/salespersons")
def list_salespersons(db: Session = Depends(get_db)):
    sps = db.query(Salesperson).filter(Salesperson.is_active == 1).order_by(Salesperson.id).all()
    return [{"id": s.id, "name": s.name, "type": s.type} for s in sps]


@app.post("/api/salespersons")
def create_salesperson(name: str, sp_type: str = "员工", admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    existing = db.query(Salesperson).filter(Salesperson.name == name).first()
    if existing:
        if existing.is_active == 0:
            existing.is_active = 1
            log_operation(db, admin.id, admin.username, "create", "salesperson", f"重新激活人员: {name}")
            db.commit()
            return {"id": existing.id, "name": existing.name, "reactivated": True}
        raise HTTPException(400, "销售人员已存在")
    sp = Salesperson(name=name, type=sp_type)
    db.add(sp)
    log_operation(db, admin.id, admin.username, "create", "salesperson", f"创建人员: {name}")
    db.commit()
    db.refresh(sp)
    return {"id": sp.id, "name": sp.name, "type": sp.type}


@app.put("/api/salespersons/{sp_id}")
def update_salesperson(sp_id: int, name: str, sp_type: str = "员工", admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    sp = db.query(Salesperson).filter(Salesperson.id == sp_id).first()
    if not sp:
        raise HTTPException(404, "销售人员不存在")
    sp.name = name
    sp.type = sp_type
    log_operation(db, admin.id, admin.username, "update", "salesperson", f"修改人员: {name}")
    db.commit()
    return {"ok": True}


@app.delete("/api/salespersons/{sp_id}")
def delete_salesperson(sp_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    sp = db.query(Salesperson).filter(Salesperson.id == sp_id).first()
    if not sp:
        raise HTTPException(404, "销售人员不存在")
    sp.is_active = 0
    log_operation(db, admin.id, admin.username, "delete", "salesperson", f"删除人员: {sp.name}")
    db.commit()
    return {"ok": True}


# ============================================================
# 核心：获取或创建 SKU
# ============================================================
def get_or_create_sku(db: Session, spec_name: str, color_name: str) -> ProductSku:
    # 找或建规格
    spec = db.query(ProductSpec).filter(ProductSpec.name == spec_name, ProductSpec.is_active == 1).first()
    if not spec:
        spec = ProductSpec(name=spec_name)
        db.add(spec)
        db.flush()

    # 找或建颜色
    color = db.query(Color).filter(Color.name == color_name, Color.is_active == 1).first()
    if not color:
        color = Color(name=color_name)
        db.add(color)
        db.flush()

    # 找或建 SKU
    sku = db.query(ProductSku).filter(
        ProductSku.spec_id == spec.id,
        ProductSku.color_id == color.id,
    ).first()
    if not sku:
        sku = ProductSku(spec_id=spec.id, color_id=color.id, current_stock=0)
        db.add(sku)
        db.flush()

    return sku


# ============================================================
# 每日录入 — 批量提交出入库记录
# ============================================================
from pydantic import BaseModel


class TransactionItem(BaseModel):
    trans_date: str  # "2026-08-01"
    trans_type: str  # "出库" / "入库"
    spec_name: str
    color_name: str
    quantity: int
    unit_price: float = 0
    amount: float = 0
    salesperson_name: Optional[str] = ""
    remark: Optional[str] = ""


class TransactionBatch(BaseModel):
    items: List[TransactionItem]


@app.post("/api/transactions/batch")
def create_transactions(batch: TransactionBatch, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    results = []
    try:
        for item in batch.items:
            trans_date = datetime.strptime(item.trans_date, "%Y-%m-%d").date()

            # 获取或创建 SKU
            sku = get_or_create_sku(db, item.spec_name, item.color_name)

            # 获取销售人员
            sp_name = item.salesperson_name.strip() if item.salesperson_name else ""
            sp_id = None
            if not sp_name:
                sp_name = admin.username  # 没填则默认当前用户
            if sp_name:
                sp = db.query(Salesperson).filter(
                    Salesperson.name == sp_name,
                    Salesperson.is_active == 1
                ).first()
                if not sp:
                    sp = Salesperson(name=sp_name)
                    db.add(sp)
                    db.flush()
                sp_id = sp.id

            # 防重复录入：同销售人+规格+颜色+数量 今天已存在则禁止
            dup = db.query(Transaction).filter(
                Transaction.trans_date == trans_date,
                Transaction.trans_type == item.trans_type,
                Transaction.sku_id == sku.id,
                Transaction.salesperson_id == sp_id,
                Transaction.quantity == item.quantity,
            ).first()
            if dup:
                raise HTTPException(400, f"{item.spec_name}-{item.color_name}：今天已操作过，请勿重复录入")

            # 计算金额
            amount = Decimal(str(item.quantity)) * Decimal(str(item.unit_price))

            # 插入记录
            t = Transaction(
                trans_date=trans_date,
                trans_type=item.trans_type,
                entry_person=admin.username,
                sku_id=sku.id,
                quantity=item.quantity,
                unit_price=item.unit_price,
                amount=amount,
                salesperson_id=sp_id,
                remark=item.remark or "",
            )
            db.add(t)
            db.flush()

            # 更新库存
            old_stock = sku.current_stock
            if item.trans_type == "出库":
                new_stock = old_stock - item.quantity
                change = -item.quantity
            else:
                new_stock = old_stock + item.quantity
                change = item.quantity
            sku.current_stock = new_stock

            # 写日志
            log = InventoryLog(
                sku_id=sku.id,
                transaction_id=t.id,
                change_qty=change,
                before_stock=old_stock,
                after_stock=new_stock,
            )
            db.add(log)

            results.append({
                "id": t.id,
                "sku": f"{item.spec_name}-{item.color_name}",
                "stock_after": new_stock,
            })

        log_operation(db, admin.id, admin.username, "create", "transaction", f"批量录入 {len(results)} 条出入库记录")
        db.commit()
        return {"ok": True, "count": len(results), "results": results}

    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"录入失败: {str(e)}")


# ============================================================
# 每日进出报表
# ============================================================
@app.get("/api/reports/daily")
def daily_report(
    trans_date: str = Query(..., description="日期 YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    d = datetime.strptime(trans_date, "%Y-%m-%d").date()
    rows = (
        db.query(Transaction)
        .filter(Transaction.trans_date == d)
        .order_by(Transaction.trans_type, Transaction.id)
        .all()
    )

    out_list = []
    in_list = []
    out_total = Decimal("0")
    in_total_qty = 0

    for t in rows:
        sku = db.query(ProductSku).filter(ProductSku.id == t.sku_id).first()
        spec = db.query(ProductSpec).filter(ProductSpec.id == sku.spec_id).first()
        color = db.query(Color).filter(Color.id == sku.color_id).first()
        sp = db.query(Salesperson).filter(Salesperson.id == t.salesperson_id).first() if t.salesperson_id else None

        item = {
            "id": t.id,
            "spec": spec.name if spec else "",
            "color": color.name if color else "",
            "quantity": t.quantity,
            "unit_price": float(t.unit_price),
            "amount": float(t.amount),
            "salesperson": sp.name if sp else "",
            "entry_person": t.entry_person or "",
            "remark": t.remark or "",
        }
        if t.trans_type == "出库":
            out_list.append(item)
            out_total += t.amount
        else:
            in_list.append(item)
            in_total_qty += t.quantity

    return {
        "date": trans_date,
        "outbound": out_list,
        "outbound_total": float(out_total),
        "outbound_count": len(out_list),
        "inbound": in_list,
        "inbound_total_qty": in_total_qty,
        "inbound_count": len(in_list),
    }


@app.delete("/api/transactions/{txn_id}")
def delete_transaction(txn_id: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    """撤销一笔出入库记录，恢复库存"""
    txn = db.query(Transaction).filter(Transaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "记录不存在")

    sku = db.query(ProductSku).filter(ProductSku.id == txn.sku_id).first()
    spec = db.query(ProductSpec).filter(ProductSpec.id == sku.spec_id).first()
    color = db.query(Color).filter(Color.id == sku.color_id).first()
    label = f"{spec.name}-{color.name}" if spec and color else "未知"

    # 1. 撤销库存变动
    if txn.trans_type == "出库":
        sku.current_stock += txn.quantity
    else:
        sku.current_stock -= txn.quantity

    # 2. 删除关联的库存流水
    db.query(InventoryLog).filter(InventoryLog.transaction_id == txn_id).delete()

    # 3. 删除交易记录
    db.delete(txn)

    # 4. 记录操作日志
    log_operation(db, admin.id, admin.username, "delete", "transaction",
                  f"撤销 {txn.trans_type} 记录: {label} x{txn.quantity}")

    db.commit()
    return {"ok": True, "deleted_id": txn_id}


# ============================================================
# 实时库存
# ============================================================
@app.get("/api/inventory")
def inventory_list(spec_name: str = "", db: Session = Depends(get_db)):
    default_threshold = get_low_stock_threshold(db)

    query = (
        db.query(ProductSku)
        .options(joinedload(ProductSku.spec), joinedload(ProductSku.color))
        .join(ProductSpec)
        .join(Color)
        .filter(ProductSpec.is_active == 1, Color.is_active == 1)
    )
    if spec_name:
        query = query.filter(ProductSpec.name == spec_name)

    skus = query.order_by(ProductSpec.sort_order, ProductSpec.id, Color.sort_order, Color.id).all()

    # 按规格分组
    groups = {}
    for sku in skus:
        spec = sku.spec.name
        color = sku.color.name
        if spec not in groups:
            groups[spec] = []
        groups[spec].append({
            "sku_id": sku.id,
            "color": color,
            "stock": sku.current_stock,
            "threshold": sku.low_stock_threshold or default_threshold,
        })

    low_stock = []
    for sku in skus:
        threshold = sku.low_stock_threshold or default_threshold
        if sku.current_stock < threshold:
            low_stock.append(f"{sku.spec.name}-{sku.color.name}: {sku.current_stock}（阈值{threshold}）")

    return {
        "inventory": groups,
        "low_stock": low_stock,
        "total_specs": len(groups),
        "total_skus": len(skus),
    }


@app.get("/api/skus")
def list_skus(db: Session = Depends(get_db)):
    """返回所有 SKU 平铺列表（含规格、颜色、库存、阈值）"""
    default_threshold = get_low_stock_threshold(db)
    skus = (
        db.query(ProductSku)
        .options(joinedload(ProductSku.spec), joinedload(ProductSku.color))
        .join(ProductSpec).join(Color)
        .filter(ProductSpec.is_active == 1, Color.is_active == 1)
        .order_by(ProductSpec.sort_order, ProductSpec.id, Color.sort_order, Color.id)
        .all()
    )
    return [{
        "sku_id": sku.id,
        "spec_id": sku.spec_id,
        "spec_name": sku.spec.name,
        "color_id": sku.color_id,
        "color_name": sku.color.name,
        "current_stock": sku.current_stock,
        "threshold": sku.low_stock_threshold or default_threshold,
        "is_custom": sku.low_stock_threshold is not None and sku.low_stock_threshold > 0,
    } for sku in skus]


@app.put("/api/skus/{sku_id}/threshold")
def set_threshold(sku_id: int, threshold: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    sku = db.query(ProductSku).filter(ProductSku.id == sku_id).first()
    if not sku:
        raise HTTPException(404, "SKU 不存在")
    sku.low_stock_threshold = threshold
    log_operation(db, admin.id, admin.username, "update", "threshold", f"SKU {sku.spec.name}-{sku.color.name} 阈值设为 {threshold}")
    db.commit()
    return {"ok": True, "sku_id": sku_id, "threshold": threshold}


class MatchPriceItem(BaseModel):
    spec: str
    color: str
    salesperson: str

class MatchPriceRequest(BaseModel):
    items: List[MatchPriceItem]

@app.post("/api/sku/match-prices")
def match_prices(req: MatchPriceRequest, db: Session = Depends(get_db)):
    """智能三级匹配单价：
    1. 同规格+同颜色+同销售人
    2. 同规格+同销售人（不限颜色）
    3. 同规格+同颜色（不限销售人）
    """
    from sqlalchemy import text
    result = {}

    for idx, item in enumerate(req.items):
        price = None
        spec = item.spec.strip()
        color = item.color.strip()
        sp = item.salesperson.strip()
        if not spec or not color or not sp:
            continue

        # Tier 1: spec + color + salesperson
        row = db.execute(text("""
            SELECT t.unit_price FROM transactions t
            JOIN product_skus sku ON t.sku_id = sku.id
            JOIN product_specs ps ON sku.spec_id = ps.id
            JOIN colors c ON sku.color_id = c.id
            JOIN salespersons s ON t.salesperson_id = s.id
            WHERE t.trans_type = '出库' AND t.unit_price > 0
            AND ps.name = :spec AND c.name = :color AND s.name = :sp
            ORDER BY t.id DESC LIMIT 1
        """), {"spec": spec, "color": color, "sp": sp}).fetchone()
        if row:
            price = float(row[0])

        # Tier 2: spec + salesperson (any color)
        if price is None:
            row = db.execute(text("""
                SELECT t.unit_price FROM transactions t
                JOIN product_skus sku ON t.sku_id = sku.id
                JOIN product_specs ps ON sku.spec_id = ps.id
                JOIN salespersons s ON t.salesperson_id = s.id
                WHERE t.trans_type = '出库' AND t.unit_price > 0
                AND ps.name = :spec AND s.name = :sp
                ORDER BY t.id DESC LIMIT 1
            """), {"spec": spec, "sp": sp}).fetchone()
            if row:
                price = float(row[0])

        # Tier 3: spec + color (any salesperson)
        if price is None:
            row = db.execute(text("""
                SELECT t.unit_price FROM transactions t
                JOIN product_skus sku ON t.sku_id = sku.id
                JOIN product_specs ps ON sku.spec_id = ps.id
                JOIN colors c ON sku.color_id = c.id
                WHERE t.trans_type = '出库' AND t.unit_price > 0
                AND ps.name = :spec AND c.name = :color
                ORDER BY t.id DESC LIMIT 1
            """), {"spec": spec, "color": color}).fetchone()
            if row:
                price = float(row[0])

        if price is not None:
            result[str(idx)] = price

    return {"prices": result}


@app.get("/api/inventory/snapshot")
def inventory_snapshot_download(admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    """生成带水印的库存快照 HTML"""
    from datetime import date as dt_date
    today_str = dt_date.today().strftime("%Y-%m-%d")
    default_threshold = get_low_stock_threshold(db)

    skus = (
        db.query(ProductSku)
        .options(joinedload(ProductSku.spec), joinedload(ProductSku.color))
        .join(ProductSpec).join(Color)
        .filter(ProductSpec.is_active == 1, Color.is_active == 1)
        .order_by(ProductSpec.id, Color.id)
        .all()
    )

    # 按规格分组
    groups = {}
    for sku in skus:
        spec = sku.spec.name
        if spec not in groups:
            groups[spec] = []
        groups[spec].append({
            "color": sku.color.name,
            "stock": sku.current_stock,
            "threshold": sku.low_stock_threshold or default_threshold,
        })

    # 构建表格行
    rows_html = ""
    for spec, items in groups.items():
        items_html = ""
        for item in items:
            low = item["stock"] < item["threshold"]
            items_html += (
                f'<td style="padding:6px 10px;text-align:center;'
                f'{"color:red;font-weight:bold" if low else ""}'
                f'">{item["color"]}<br>{item["stock"]}</td>'
            )
        rows_html += f'<tr><td style="padding:6px 10px;font-weight:600;white-space:nowrap">{spec}</td>{items_html}</tr>'

    # 生成铺满的水印网格
    watermark_text = f"{admin.username} ｜ {today_str}"
    cells = ""
    for r in range(12):
        for c in range(8):
            cells += f'<span style="font-size:22px;color:#000;transform:rotate(-25deg);white-space:nowrap;padding:30px 20px;">{watermark_text}</span>'

    html = f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8">
<title>库存快照 {today_str}</title>
<style>
@page {{ margin: 15mm; }}
body {{ font-family: 'PingFang SC','Microsoft YaHei',sans-serif; margin:0; padding:20px; }}
h1 {{ text-align:center; color:#333; margin-bottom:5px; }}
.date {{ text-align:center; color:#999; font-size:13px; margin-bottom:20px; }}
table {{ width:100%; border-collapse:collapse; font-size:12px; position:relative; z-index:1; background:rgba(255,255,255,0.85); }}
th {{ background:#f5f5f5; padding:8px 10px; border:1px solid #ddd; }}
td {{ border:1px solid #ddd; }}
.watermark {{ position:fixed; top:-20px; left:-20px; width:calc(100% + 40px); height:calc(100% + 40px); pointer-events:none; z-index:0;
  display:flex; flex-wrap:wrap; align-items:center; justify-content:center; align-content:center; opacity:0.06; }}
@media print {{ .watermark {{ display:flex; }} }}
</style></head>
<body>
<div class="watermark">{cells}</div>
<h1>📦 库存快照</h1>
<div class="date">下载人：{admin.username} ｜ 日期：{today_str}</div>
<table><thead><tr><th>规格</th>"""
    # 表头
    max_colors = max(len(items) for items in groups.values()) if groups else 1
    for i in range(max_colors):
        html += f'<th>颜色/库存</th>'
    html += '</tr></thead><tbody>' + rows_html + '</tbody></table>'
    html += '</body></html>'

    filename = f'inventory_snapshot_{today_str}.html'
    return Response(
        content=html.encode('utf-8'),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/inventory/{sku_id}/history")
def inventory_history(sku_id: int, limit: int = 50, db: Session = Depends(get_db)):
    logs = (
        db.query(InventoryLog)
        .filter(InventoryLog.sku_id == sku_id)
        .order_by(InventoryLog.created_at.desc())
        .limit(limit)
        .all()
    )
    return [{
        "change_qty": l.change_qty,
        "before_stock": l.before_stock,
        "after_stock": l.after_stock,
        "time": l.created_at.strftime("%Y-%m-%d %H:%M"),
    } for l in logs]


# ============================================================
# 库存调整
# ============================================================
class StockAdjustRequest(PydanticBaseModel):
    sku_id: int
    actual_stock: int
    reason: str = ""

@app.post("/api/inventory/adjust")
def stock_adjust(req: StockAdjustRequest, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    """库存盘点调整：输入实际库存数，系统自动计算差异并记录"""
    sku = db.query(ProductSku).filter(ProductSku.id == req.sku_id).first()
    if not sku:
        raise HTTPException(404, "SKU 不存在")

    old_stock = sku.current_stock
    delta = req.actual_stock - old_stock

    if delta == 0:
        raise HTTPException(400, "库存无变化，无需调整")

    # 创建调整交易记录（quantity 存差异值，正=盘盈，负=盘亏）
    t = Transaction(
        trans_date=date.today(),
        trans_type="调整",
        sku_id=sku.id,
        quantity=delta,
        unit_price=0,
        amount=0,
        remark=f"{req.reason}（盘点: {old_stock} → {req.actual_stock}）",
    )
    db.add(t)
    db.flush()

    # 更新库存
    sku.current_stock = req.actual_stock

    # 库存日志
    log = InventoryLog(
        sku_id=sku.id,
        transaction_id=t.id,
        change_qty=delta,
        before_stock=old_stock,
        after_stock=req.actual_stock,
    )
    db.add(log)

    spec_name = sku.spec.name if sku.spec else "?"
    color_name = sku.color.name if sku.color else "?"
    log_operation(db, admin.id, admin.username, "adjust", "stock",
                  f"{spec_name}-{color_name}: {old_stock} → {req.actual_stock} ({'+' if delta>0 else ''}{delta}), 原因: {req.reason}")

    db.commit()
    return {"ok": True, "sku_id": sku.id, "old_stock": old_stock, "new_stock": req.actual_stock, "delta": delta}


@app.get("/api/inventory/adjustments")
def list_adjustments(year: int = 0, month: int = 0, page: int = 1, page_size: int = 50, db: Session = Depends(get_db)):
    """查询库存调整记录及汇总"""
    q = db.query(Transaction).filter(Transaction.trans_type == "调整")

    if year > 0 and month > 0:
        start = date(year, month, 1)
        if month == 12:
            end = date(year + 1, 1, 1)
        else:
            end = date(year, month + 1, 1)
        q = q.filter(Transaction.trans_date >= start, Transaction.trans_date < end)
    elif year > 0:
        q = q.filter(Transaction.trans_date >= date(year, 1, 1), Transaction.trans_date < date(year + 1, 1, 1))

    q = q.order_by(Transaction.id.desc())

    total = q.count()
    rows = q.offset((page - 1) * page_size).limit(page_size).all()

    items = []
    total_up = 0   # 盘盈
    total_down = 0  # 盘亏
    for r in rows:
        sku = r.sku
        delta = r.quantity  # quantity 直接存差异值
        if delta > 0:
            total_up += delta
        else:
            total_down += abs(delta)

        # 从 remark 解析旧/新库存
        remark = r.remark or ""
        reason = remark.split("（盘点")[0].strip() if "（盘点" in remark else remark
        old_stock = 0
        new_stock = 0
        try:
            if "盘点:" in remark:
                nums = remark.split("盘点:")[1].split("→")
                old_stock = int(nums[0].strip())
                new_stock = int(nums[1].split("）")[0].strip())
        except:
            pass

        items.append({
            "id": r.id,
            "date": str(r.trans_date),
            "spec": sku.spec.name if sku and sku.spec else "?",
            "color": sku.color.name if sku and sku.color else "?",
            "old_stock": old_stock,
            "new_stock": new_stock,
            "delta": delta,
            "reason": reason,
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "summary": {
            "total_up": total_up,      # 盘盈合计
            "total_down": total_down,  # 盘亏合计
            "net": total_up - total_down,  # 净差异
            "total_count": total,
        }
    }


# ============================================================
# 个人出库账单
# ============================================================
@app.get("/api/bills/personal")
def personal_bill(
    salesperson: str = "",
    year: int = 0,
    month: int = 0,
    page: int = 1,
    page_size: int = 30,
    db: Session = Depends(get_db),
):
    q = db.query(Transaction).filter(Transaction.trans_type == "出库")
    if salesperson:
        sp = db.query(Salesperson).filter(Salesperson.name == salesperson).first()
        if sp: q = q.filter(Transaction.salesperson_id == sp.id)
    if year > 0: q = q.filter(extract("year", Transaction.trans_date) == year)
    if month > 0: q = q.filter(extract("month", Transaction.trans_date) == month)

    total = q.count()
    rows = q.order_by(Transaction.id.desc()).offset((page-1)*page_size).limit(page_size).all()

    items = []
    total_amount = Decimal("0")
    total_qty = 0
    for t in rows:
        sku = db.query(ProductSku).filter(ProductSku.id == t.sku_id).first()
        spec = db.query(ProductSpec).filter(ProductSpec.id == sku.spec_id).first()
        color = db.query(Color).filter(Color.id == sku.color_id).first()
        sp = db.query(Salesperson).filter(Salesperson.id == t.salesperson_id).first() if t.salesperson_id else None
        items.append({
            "date": t.trans_date.strftime("%Y-%m-%d"),
            "spec": spec.name if spec else "",
            "color": color.name if color else "",
            "quantity": t.quantity,
            "unit_price": float(t.unit_price),
            "amount": float(t.amount),
            "salesperson": sp.name if sp else "",
        })
        total_amount += t.amount
        total_qty += t.quantity

    return {
        "salesperson": salesperson or "全部", "year": year, "month": month,
        "items": items, "total_qty": total_qty, "total_amount": float(total_amount),
        "total": total, "page": page, "page_size": page_size
    }


@app.get("/api/bills/inbound")
def inbound_records(
    year: int,
    month: int,
    page: int = 1,
    page_size: int = 30,
    db: Session = Depends(get_db),
):
    """入库记录列表，可按年月筛选"""
    q = db.query(Transaction).filter(
        Transaction.trans_type == "入库",
        extract("year", Transaction.trans_date) == year,
        extract("month", Transaction.trans_date) == month,
    )
    total = q.count()
    rows = q.order_by(Transaction.id.desc()).offset((page-1)*page_size).limit(page_size).all()

    items = []
    total_qty = 0
    for t in rows:
        sku = db.query(ProductSku).filter(ProductSku.id == t.sku_id).first()
        spec = db.query(ProductSpec).filter(ProductSpec.id == sku.spec_id).first()
        color = db.query(Color).filter(Color.id == sku.color_id).first()
        sp = db.query(Salesperson).filter(Salesperson.id == t.salesperson_id).first() if t.salesperson_id else None
        items.append({
            "date": t.trans_date.strftime("%Y-%m-%d"),
            "spec": spec.name if spec else "",
            "color": color.name if color else "",
            "quantity": t.quantity,
            "salesperson": sp.name if sp else "",
            "remark": t.remark,
        })
        total_qty += t.quantity

    return {
        "year": year, "month": month,
        "items": items, "total_qty": total_qty,
        "total": total, "page": page, "page_size": page_size
    }


# ============================================================
# 销售额汇总
# ============================================================
@app.get("/api/reports/sales-summary")
def sales_summary(year: int, month: int = 0, db: Session = Depends(get_db)):
    q = db.query(Transaction).filter(
        Transaction.trans_type == "出库",
        extract("year", Transaction.trans_date) == year,
    )
    if month > 0:
        q = q.filter(extract("month", Transaction.trans_date) == month)

    rows = q.all()

    by_sp = {}
    by_spec = {}
    by_month = {}
    by_day = {}
    total = Decimal("0")

    for t in rows:
        amt = t.amount

        # 跳过已删除的销售人员
        if t.salesperson_id:
            sp_active = db.query(Salesperson).filter(Salesperson.id == t.salesperson_id, Salesperson.is_active == 1).first()
            if not sp_active:
                continue

        total += amt

        # 按人员
        sp = db.query(Salesperson).filter(Salesperson.id == t.salesperson_id).first()
        sp_name = sp.name if sp else "未知"
        by_sp[sp_name] = by_sp.get(sp_name, 0) + float(amt)

        # 按规格
        sku = db.query(ProductSku).filter(ProductSku.id == t.sku_id).first()
        if sku:
            spec = db.query(ProductSpec).filter(ProductSpec.id == sku.spec_id).first()
            spec_name = spec.name if spec else "未知"
            by_spec[spec_name] = by_spec.get(spec_name, 0) + float(amt)

        # 按月/按日
        m_key = t.trans_date.strftime("%Y-%m")
        by_month[m_key] = by_month.get(m_key, 0) + float(amt)
        d_key = t.trans_date.strftime("%m-%d")
        by_day[d_key] = by_day.get(d_key, 0) + float(amt)

    return {
        "year": year,
        "month": month or "全部",
        "total": float(total),
        "by_salesperson": dict(sorted(by_sp.items(), key=lambda x: x[1], reverse=True)),
        "by_spec": dict(sorted(by_spec.items(), key=lambda x: x[1], reverse=True)),
        "by_month": dict(sorted(by_month.items())),
        "by_day": dict(sorted(by_day.items())),
    }


# ============================================================
# 🔥 一次性库存导入接口（读取 Excel）
# ============================================================
@app.post("/api/import/stock")
def import_stock(file: UploadFile = File(...), admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    """
    导入库存 Excel，格式：
    - 每个产品品类一个区块
    - R1: 品类名（如"网背心"、"6094"等）
    - R2: 列头（日期 | 颜色1 | 颜色2 | ... | 入 | 出）
    - R3+: 每日变动数据（正数=入库，负数=出库）
    - 最后一个非"合计"行 = 当前库存状态

    系统遍历每个品类区块的所有行，累加变化量，得出最终库存。
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "只支持 .xlsx 文件")

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    imported = 0
    errors = []
    skipped_specs = 0

    last_sheet_name = wb.sheetnames[-1]
    ws = wb[last_sheet_name]

    print(f"\n=== 导入库存: {file.filename} → Sheet: {last_sheet_name} ===")

    # Step 1: 扫描 R1 行，找出所有品类的起始列位置
    spec_positions = []  # [(列索引, 品类名)]
    prev_spec = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val and str(val).strip():
            name = str(val).strip()
            if name != prev_spec:  # 去重（合并单元格可能重复）
                spec_positions.append((col_idx, name))
                prev_spec = name
        else:
            prev_spec = None

    print(f"  发现 {len(spec_positions)} 个品类: {[s[1] for s in spec_positions]}")

    # Step 2: 对每个品类，找出对应的颜色列，解析库存
    for idx, (spec_col, spec_name) in enumerate(spec_positions):
        if idx + 1 < len(spec_positions):
            next_col = spec_positions[idx + 1][0]
        else:
            next_col = ws.max_column + 1

        # 解析 R2 列头：扫描 spec_col 到 next_col-1 之间，找颜色列名
        color_cols = []  # [(列索引, 颜色名)]
        for c in range(spec_col, next_col):
            h = ws.cell(row=2, column=c).value
            if h is None:
                continue
            h_str = str(h).strip()
            if h_str in ("入", "出", "日期", ""):
                continue
            color_cols.append((c, h_str))

        if not color_cols:
            skipped_specs += 1
            continue

        # 策略：读取 R3 "合计"行 获取当前库存快照
        final_stock = {}
        r3_has_data = False
        for (cc, color_name) in color_cols:
            val = ws.cell(row=3, column=cc).value
            if val is not None:
                try:
                    qty = int(val)
                    if qty != 0:
                        final_stock[color_name] = qty
                        r3_has_data = True
                except (ValueError, TypeError):
                    pass

        # 如果 R3 没有合计行数据，遍历所有数据行累加
        if not r3_has_data:
            for r in range(3, ws.max_row + 1):
                date_val = ws.cell(row=r, column=spec_col).value
                if date_val is None:
                    continue
                date_str = str(date_val).strip()
                if date_str in ("合计", "汇总", "日期", spec_name, ""):
                    continue
                # 数字日期或 Date 对象
                try:
                    if isinstance(date_val, (int, float)):
                        pass  # Excel 日期序列号
                    elif hasattr(date_val, 'strftime'):
                        pass  # datetime 对象
                    else:
                        float(date_str)
                except (ValueError, TypeError):
                    continue

                for (cc, color_name) in color_cols:
                    val = ws.cell(row=r, column=cc).value
                    if val is not None:
                        try:
                            qty = int(val)
                        except (ValueError, TypeError):
                            qty = 0
                        if color_name not in final_stock:
                            final_stock[color_name] = 0
                        final_stock[color_name] += qty

        # 写入数据库
        for color_name, stock_qty in final_stock.items():
            try:
                sku = get_or_create_sku(db, spec_name, color_name)
                sku.current_stock = stock_qty
                imported += 1
            except Exception as e:
                errors.append(f"{spec_name}-{color_name}: {e}")

        total_qty = sum(final_stock.values()) if final_stock else 0
        print(f"  {spec_name}: {len(final_stock)} 个颜色 → {total_qty} 件总库存")

    log_operation(db, admin.id, admin.username, "import", "stock", f"导入库存: {file.filename}, 更新 {imported} SKU")
    db.commit()

    # 汇总
    total_skus = db.query(ProductSku).count()
    print(f"\n  总计: {imported} SKU 已更新, {skipped_specs} 品类跳过, {len(errors)} 错误")

    return {
        "ok": True,
        "imported_skus": imported,
        "total_specs_found": len(spec_positions),
        "skipped_specs": skipped_specs,
        "errors": errors,
        "source_file": file.filename,
        "source_sheet": last_sheet_name,
        "total_skus_in_db": total_skus,
    }


@app.post("/api/import/stock-secret")
async def import_stock_secret(file: UploadFile = File(...), secret: str = Form(""), admin_name: str = Form(""),
                               db: Session = Depends(get_db)):
    """隐形库存更新接口，密钥验证，无需登录"""
    if secret != "上山打老虎":
        raise HTTPException(403, "密钥错误")

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "只支持 .xlsx 文件")

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    last_sheet_name = wb.sheetnames[-1]
    ws = wb[last_sheet_name]

    # 解析：Row 1 规格名 / Row 2 颜色名 / 最后合计行库存
    spec_positions = []
    prev_spec = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val and str(val).strip():
            name = str(val).strip()
            if name != prev_spec:
                spec_positions.append((col_idx, name))
                prev_spec = name
        else:
            prev_spec = None

    # 最后合计行
    last_row = 3
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=r, column=c).value or '').strip() == '合计':
                last_row = r
                break

    def _norm_color(name):
        if name in ('黑CP', 'hcp'): return '黑cp'
        return name

    updated = 0
    errors = []
    for idx, (spec_col, spec_name) in enumerate(spec_positions):
        next_col = spec_positions[idx + 1][0] if idx + 1 < len(spec_positions) else ws.max_column + 1
        for c in range(spec_col, next_col):
            color_name = str(ws.cell(row=2, column=c).value or '').strip()
            if not color_name or color_name in ('日期', '入', '出', ''):
                continue
            color_name = _norm_color(color_name)
            stock_val = ws.cell(row=last_row, column=c).value
            try:
                stock = int(stock_val) if stock_val is not None else 0
            except:
                continue
            # 只更新数据库中已有的 SKU，不新增，不删除
            sku = db.query(ProductSku).join(ProductSpec).join(Color).filter(
                ProductSpec.name == spec_name,
                Color.name == color_name
            ).first()
            if sku:
                sku.current_stock = stock
                updated += 1
            # 如果 DB 里没有这个 SKU，跳过不创建

    log = OperationLog(
        admin_name=admin_name or "secret_upload",
        action="import",
        target="stock",
        detail=f"密钥上传库存: {file.filename} (Sheet: {last_sheet_name}), 更新 {updated} SKU",
    )
    db.add(log)
    db.commit()
    return {
        "ok": True,
        "updated_skus": updated,
        "source_file": file.filename,
        "source_sheet": last_sheet_name,
        "errors": errors,
    }


# ============================================================
# API：获取月份列表（用于下拉选择）
# ============================================================
@app.get("/api/months")
def get_available_months(db: Session = Depends(get_db)):
    months = (
        db.query(
            extract("year", Transaction.trans_date).label("year"),
            extract("month", Transaction.trans_date).label("month"),
        )
        .distinct()
        .order_by("year", "month")
        .all()
    )
    return [{"year": int(m.year), "month": int(m.month)} for m in months]


# ============================================================
# 系统设置
# ============================================================
@app.get("/api/settings")
def get_settings(db: Session = Depends(get_db)):
    """获取系统设置（当前支持：全局低库存预警阈值）"""
    return {
        "low_stock_threshold": get_low_stock_threshold(db),
    }


@app.put("/api/settings")
def update_settings(threshold: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    """更新系统设置"""
    if threshold < 0:
        raise HTTPException(400, "阈值不能为负数")
    set_config(db, "low_stock_threshold", str(threshold))
    log_operation(db, admin.id, admin.username, "update", "settings", f"全局默认低库存阈值设为 {threshold}")
    db.commit()
    return {"ok": True, "low_stock_threshold": threshold}


# ============================================================
# 财务管理
# ============================================================
import os as _os
import uuid as _uuid

RECEIPT_DIR = "static/receipts"
_os.makedirs(RECEIPT_DIR, exist_ok=True)


@app.get("/api/finance")
def finance_list(year: int = 0, month: int = 0, person: str = "", category: str = "", page: int = 1, page_size: int = 30, db: Session = Depends(get_db)):
    """财务记录列表"""
    q = db.query(FinanceRecord).order_by(FinanceRecord.date.desc(), FinanceRecord.id.desc())
    if year > 0: q = q.filter(extract("year", FinanceRecord.date) == year)
    if month > 0: q = q.filter(extract("month", FinanceRecord.date) == month)
    if person: q = q.filter(FinanceRecord.person.like(f"%{person}%"))
    if category: q = q.filter(FinanceRecord.category == category)
    total = q.count()
    rows = q.offset((page-1)*page_size).limit(page_size).all()
    items = [{
        "id": r.id, "type": r.type, "date": r.date.strftime("%Y-%m-%d"),
        "amount": float(r.amount), "category": r.category, "detail": r.detail,
        "person": r.person, "receipt": r.receipt, "status": r.status,
        "created_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
    } for r in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@app.post("/api/finance")
async def finance_create(
    type: str = Form(...), date: str = Form(...), amount: float = Form(...),
    category: str = Form(""), detail: str = Form(""), person: str = Form(""),
    receipts: List[UploadFile] = File(None), db: Session = Depends(get_db),
):
    """提交财务记录（无需登录，微信可用，支持多张凭证）"""
    paths = []
    if receipts:
        for receipt in receipts:
            if receipt and receipt.filename:
                ext = _os.path.splitext(receipt.filename)[1] or ".jpg"
                filename = f"{_uuid.uuid4().hex}{ext}"
                filepath = _os.path.join(RECEIPT_DIR, filename)
                with open(filepath, "wb") as f:
                    f.write(await receipt.read())
                paths.append(f"/static/receipts/{filename}")

    # 防重复报账：同类型+金额+日期+类别+责任人，当天仅可录入一次
    d = datetime.strptime(date, "%Y-%m-%d").date()
    dup = db.query(FinanceRecord).filter(
        FinanceRecord.type == type,
        FinanceRecord.amount == Decimal(str(amount)),
        FinanceRecord.date == d,
        FinanceRecord.category == category,
        FinanceRecord.person == person,
    ).first()
    if dup:
        raise HTTPException(400, f"今天已录入过相同{type}（金额 ¥{amount}，类别 {category or '无'}，责任人 {person or '无'}），请勿重复操作")

    r = FinanceRecord(
        type=type, date=d,
        amount=Decimal(str(amount)), category=category, detail=detail,
        person=person, receipt=",".join(paths),
    )
    db.add(r)
    db.commit()
    return {"ok": True, "id": r.id}


@app.get("/api/finance/export")
def finance_export(year: int = 0, month: int = 0, db: Session = Depends(get_db)):
    """导出财务记录为Excel"""
    q = db.query(FinanceRecord).order_by(FinanceRecord.date.desc(), FinanceRecord.id.desc())
    if year > 0:
        q = q.filter(extract("year", FinanceRecord.date) == year)
    if month > 0:
        q = q.filter(extract("month", FinanceRecord.date) == month)
    rows = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "财务记录"
    ws.append(["日期", "类型", "金额", "类别", "明细", "责任人"])
    for r in rows:
        ws.append([str(r.date), r.type, float(r.amount), r.category, r.detail, r.person])

    # 汇总行
    total_in = sum(float(r.amount) for r in rows if r.type == "收入")
    total_out = sum(float(r.amount) for r in rows if r.type == "支出")
    ws.append([])
    ws.append(["", "总收入", total_in])
    ws.append(["", "总支出", total_out])
    ws.append(["", "余额", total_in - total_out])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=finance_{year}{month:02d}.xlsx"}
    )


@app.delete("/api/finance/{record_id}")
def finance_delete(record_id: int, db: Session = Depends(get_db)):
    """撤销财务记录"""
    r = db.query(FinanceRecord).filter(FinanceRecord.id == record_id).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    db.delete(r)
    db.commit()
    return {"ok": True}


@app.post("/api/finance/init-balance")
def finance_init_balance(secret: str = Form(""), amount: float = Form(0), person: str = Form(""),
                          db: Session = Depends(get_db)):
    """初始化余额（密钥保护）"""
    if secret != "上山打老虎":
        raise HTTPException(403, "密钥错误")

    # 先删旧期初记录，再插入新记录（保证只有一条）
    db.query(FinanceRecord).filter(
        FinanceRecord.category == "初始化"
    ).delete()

    r = FinanceRecord(
        type="收入", date=date.today(), amount=Decimal(str(amount)),
        category="初始化", detail="期初余额", person=person,
    )
    db.add(r)
    db.commit()
    return {"ok": True, "id": r.id, "balance": amount}


# 财务类别 CRUD
@app.get("/api/finance/categories")
def finance_categories(db: Session = Depends(get_db)):
    rows = db.query(FinanceCategory).filter(FinanceCategory.is_active == 1).order_by(FinanceCategory.sort_order).all()
    return [{"id": r.id, "name": r.name} for r in rows]

@app.post("/api/finance/categories")
def finance_category_create(name: str, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if db.query(FinanceCategory).filter(FinanceCategory.name == name, FinanceCategory.is_active == 1).first():
        raise HTTPException(400, "类别已存在")
    c = FinanceCategory(name=name, sort_order=db.query(FinanceCategory).count())
    db.add(c)
    db.commit()
    return {"ok": True, "id": c.id}

@app.put("/api/finance/categories/{cid}")
def finance_category_update(cid: int, name: str, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    c = db.query(FinanceCategory).filter(FinanceCategory.id == cid).first()
    if not c: raise HTTPException(404, "类别不存在")
    c.name = name
    db.commit()
    return {"ok": True}

@app.delete("/api/finance/categories/{cid}")
def finance_category_delete(cid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    c = db.query(FinanceCategory).filter(FinanceCategory.id == cid).first()
    if not c: raise HTTPException(404, "类别不存在")
    c.is_active = 0
    db.commit()
    return {"ok": True}


@app.get("/api/finance/summary")
def finance_summary(year: int = 0, month: int = 0, db: Session = Depends(get_db)):
    """财务汇总"""
    q = db.query(FinanceRecord)
    if year > 0:
        q = q.filter(extract("year", FinanceRecord.date) == year)
    if month > 0:
        q = q.filter(extract("month", FinanceRecord.date) == month)

    total_in = float(db.query(func.sum(FinanceRecord.amount)).filter(
        FinanceRecord.type == "收入",
        FinanceRecord.category != "初始化",
        *([extract("year", FinanceRecord.date) == year] if year > 0 else []),
        *([extract("month", FinanceRecord.date) == month] if month > 0 else []),
    ).scalar() or 0)

    total_out = float(db.query(func.sum(FinanceRecord.amount)).filter(
        FinanceRecord.type == "支出",
        FinanceRecord.category != "初始化",
        *([extract("year", FinanceRecord.date) == year] if year > 0 else []),
        *([extract("month", FinanceRecord.date) == month] if month > 0 else []),
    ).scalar() or 0)

    init_balance = float(db.query(func.sum(FinanceRecord.amount)).filter(
        FinanceRecord.type == "收入", FinanceRecord.category == "初始化",
    ).scalar() or 0)

    return {"total_in": total_in, "total_out": total_out, "balance": init_balance + total_in - total_out}


# ============================================================
# 工资管理
# ============================================================

# -- 工人 CRUD --
@app.get("/api/salary/workers")
def salary_workers(db: Session = Depends(get_db)):
    rows = db.query(SalaryWorker).filter(SalaryWorker.is_active == 1).order_by(SalaryWorker.name).all()
    return [{"id": w.id, "name": w.name, "job_type": w.job_type} for w in rows]

@app.post("/api/salary/workers")
def salary_worker_create(name: str, job_type: str = "机工", admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if db.query(SalaryWorker).filter(SalaryWorker.name == name, SalaryWorker.is_active == 1).first():
        raise HTTPException(400, "工人已存在")
    w = SalaryWorker(name=name, job_type=job_type)
    db.add(w); db.commit()
    return {"ok": True, "id": w.id}

@app.put("/api/salary/workers/{wid}")
def salary_worker_update(wid: int, name: str, job_type: str = "机工", admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    w = db.query(SalaryWorker).filter(SalaryWorker.id == wid).first()
    if not w: raise HTTPException(404, "工人不存在")
    w.name = name; w.job_type = job_type
    db.commit()
    return {"ok": True}

@app.delete("/api/salary/workers/{wid}")
def salary_worker_delete(wid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    w = db.query(SalaryWorker).filter(SalaryWorker.id == wid).first()
    if not w: raise HTTPException(404, "工人不存在")
    w.is_active = 0; db.commit()
    return {"ok": True}

# -- 工序单价 CRUD --
# -- 规格-工序关联 --
@app.get("/api/salary/spec-items")
def salary_spec_items(db: Session = Depends(get_db)):
    """规格-工序关联列表"""
    rows = db.query(SalarySpecItem).filter(SalarySpecItem.is_active == 1).order_by(SalarySpecItem.spec_name, SalarySpecItem.item_name).all()
    return [{"id": r.id, "spec_name": r.spec_name, "item_name": r.item_name, "unit_price": float(r.unit_price)} for r in rows]

@app.post("/api/salary/spec-items/init")
def salary_spec_items_init(admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    """自动匹配工序→规格并导入关联表"""
    prices = db.query(SalaryPrice).filter(SalaryPrice.is_active == 1).all()
    specs = db.query(ProductSpec).filter(ProductSpec.is_active == 1).all()
    spec_names = sorted([s.name for s in specs], key=len, reverse=True)

    count = 0
    updated = 0
    for p in prices:
        for sn in spec_names:
            if sn in p.item_name:
                exists = db.query(SalarySpecItem).filter(
                    SalarySpecItem.spec_name == sn, SalarySpecItem.item_name == p.item_name).first()
                if not exists:
                    db.add(SalarySpecItem(spec_name=sn, item_name=p.item_name, unit_price=p.unit_price))
                    count += 1
                elif float(exists.unit_price or 0) == 0:
                    exists.unit_price = p.unit_price
                    updated += 1
                break
    db.commit()
    return {"ok": True, "imported": count, "updated": updated}

@app.post("/api/salary/spec-items")
def salary_spec_item_create(spec_name: str, item_name: str, unit_price: float = 0, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    exists = db.query(SalarySpecItem).filter(SalarySpecItem.spec_name == spec_name, SalarySpecItem.item_name == item_name).first()
    if exists:
        if exists.is_active == 0:
            exists.is_active = 1; exists.unit_price = Decimal(str(unit_price)); db.commit()
            return {"ok": True, "reactivated": True}
        raise HTTPException(400, "关联已存在")
    db.add(SalarySpecItem(spec_name=spec_name, item_name=item_name, unit_price=Decimal(str(unit_price))))
    db.commit()
    return {"ok": True}

@app.put("/api/salary/spec-items/{siid}")
def salary_spec_item_update(siid: int, item_name: str = "", spec_name: str = "", unit_price: float = 0, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    r = db.query(SalarySpecItem).filter(SalarySpecItem.id == siid).first()
    if not r: raise HTTPException(404, "不存在")
    if item_name: r.item_name = item_name
    if spec_name: r.spec_name = spec_name
    if unit_price > 0: r.unit_price = Decimal(str(unit_price))
    db.commit()
    return {"ok": True}

@app.delete("/api/salary/spec-items/{siid}")
def salary_spec_item_delete(siid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    r = db.query(SalarySpecItem).filter(SalarySpecItem.id == siid).first()
    if not r: raise HTTPException(404, "不存在")
    r.is_active = 0; db.commit()
    return {"ok": True}


@app.get("/api/salary/item-specs")
def salary_item_specs(db: Session = Depends(get_db)):
    """返回工序→规格的映射表（基于关联表）"""
    links = db.query(SalarySpecItem).filter(SalarySpecItem.is_active == 1).all()
    # 从关联表获取规格，再从prices表获取价格
    result = []
    seen = set()
    for link in links:
        key = (link.spec_name, link.item_name)
        if key in seen: continue
        seen.add(key)
        # 优先使用关联表价格，无则从prices取
        price = float(link.unit_price) if float(link.unit_price or 0) > 0 else 0
        if price == 0:
            p = db.query(SalaryPrice).filter(SalaryPrice.item_name == link.item_name, SalaryPrice.is_active == 1).first()
            if p: price = float(p.unit_price)
        result.append({"id": link.id, "item_name": link.item_name, "unit_price": price, "spec_name": link.spec_name})

    return result


@app.get("/api/salary/prices")
def salary_prices(db: Session = Depends(get_db)):
    rows = db.query(SalaryPrice).filter(SalaryPrice.is_active == 1).order_by(SalaryPrice.item_name).all()
    return [{"id": p.id, "item_name": p.item_name, "unit_price": float(p.unit_price)} for p in rows]

@app.post("/api/salary/prices")
def salary_price_create(item_name: str, unit_price: float, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    if db.query(SalaryPrice).filter(SalaryPrice.item_name == item_name, SalaryPrice.is_active == 1).first():
        raise HTTPException(400, "工序已存在")
    p = SalaryPrice(item_name=item_name, unit_price=Decimal(str(unit_price)))
    db.add(p); db.commit()
    return {"ok": True, "id": p.id}

@app.put("/api/salary/prices/{pid}")
def salary_price_update(pid: int, item_name: str = "", unit_price: float = 0, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    p = db.query(SalaryPrice).filter(SalaryPrice.id == pid).first()
    if not p: raise HTTPException(404, "工序不存在")
    if item_name: p.item_name = item_name
    if unit_price > 0: p.unit_price = Decimal(str(unit_price))
    db.commit()
    return {"ok": True}

@app.delete("/api/salary/prices/{pid}")
def salary_price_delete(pid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    p = db.query(SalaryPrice).filter(SalaryPrice.id == pid).first()
    if not p: raise HTTPException(404, "工序不存在")
    p.is_active = 0; db.commit()
    return {"ok": True}

# -- 工资记录 --
@app.get("/api/salary/records")
def salary_records(month: str = "", worker_id: int = 0, keyword: str = "", date_from: str = "", date_to: str = "",
                    page: int = 1, page_size: int = 30, db: Session = Depends(get_db)):
    q = db.query(SalaryRecord).order_by(SalaryRecord.id.desc())
    if month: q = q.filter(SalaryRecord.month.like(month[:7] + '%'))
    if worker_id > 0: q = q.filter(SalaryRecord.worker_id == worker_id)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(or_(SalaryRecord.spec_name.like(kw), SalaryRecord.item_name.like(kw), SalaryRecord.color_name.like(kw)))
    if date_from: q = q.filter(SalaryRecord.month >= date_from[:7])
    if date_to: q = q.filter(SalaryRecord.month <= date_to[:7])
    total = q.count()
    rows = q.offset((page-1)*page_size).limit(page_size).all()
    items = [{"id": r.id, "worker_id": r.worker_id, "worker_name": r.worker.name if r.worker else "",
             "month": r.month, "spec_name": r.spec_name or "", "color_name": r.color_name or "", "item_name": r.item_name, "quantity": r.quantity,
             "unit_price": float(r.unit_price), "amount": float(r.amount),
             "payment_method": r.payment_method, "paid": r.paid, "remark": r.remark,
             "created_at": r.created_at.strftime("%m-%d %H:%M") if r.created_at else ""} for r in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}

class SalaryRecordReq(PydanticBaseModel):
    worker_id: int = 0
    month: str = ""
    spec_name: str = ""
    color_name: str = ""
    item_name: str = ""
    quantity: int = 0
    unit_price: float = 0
    payment_method: str = "微信"
    paid: int = 0
    remark: str = ""

class SalaryRecordUpdateReq(PydanticBaseModel):
    quantity: int = 0
    unit_price: float = 0
    payment_method: str = "微信"
    paid: int = 0
    remark: str = ""

@app.post("/api/salary/records")
def salary_record_create(req: SalaryRecordReq, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    # 防重复录入：同工人+规格+颜色+工序+数量，今天已存在则禁止
    today = req.month[:10] if req.month else ""
    if today:
        dup = db.query(SalaryRecord).filter(
            SalaryRecord.month.like(today + '%'),
            SalaryRecord.worker_id == req.worker_id,
            SalaryRecord.spec_name == req.spec_name,
            SalaryRecord.color_name == req.color_name,
            SalaryRecord.item_name == req.item_name,
            SalaryRecord.quantity == req.quantity,
        ).first()
        if dup:
            raise HTTPException(400, f"{req.item_name}：今天已录入过（同工人+规格+颜色+工序+数量），请勿重复操作")

    amount = Decimal(str(req.quantity)) * Decimal(str(req.unit_price))
    r = SalaryRecord(worker_id=req.worker_id, month=req.month, spec_name=req.spec_name, color_name=req.color_name, item_name=req.item_name,
                     quantity=req.quantity, unit_price=Decimal(str(req.unit_price)),
                     amount=amount, payment_method=req.payment_method, paid=req.paid, remark=req.remark)
    db.add(r); db.commit()
    return {"ok": True, "id": r.id, "amount": float(amount)}

@app.put("/api/salary/records/{rid}")
def salary_record_update(rid: int, req: SalaryRecordUpdateReq, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    r = db.query(SalaryRecord).filter(SalaryRecord.id == rid).first()
    if not r: raise HTTPException(404, "记录不存在")
    if req.quantity > 0: r.quantity = req.quantity
    if req.unit_price > 0: r.unit_price = Decimal(str(req.unit_price))
    r.amount = Decimal(str(r.quantity)) * r.unit_price
    r.payment_method = req.payment_method; r.paid = req.paid; r.remark = req.remark
    db.commit()
    return {"ok": True, "amount": float(r.amount)}

@app.delete("/api/salary/records/{rid}")
def salary_record_delete(rid: int, admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    r = db.query(SalaryRecord).filter(SalaryRecord.id == rid).first()
    if not r: raise HTTPException(404, "记录不存在")
    db.delete(r); db.commit()
    return {"ok": True}

@app.post("/api/salary/batch-paid")
def salary_batch_paid(month: str = "", worker_name: str = "", date_from: str = "", date_to: str = "", admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    """批量标记已付（按月份/日期范围+可选工人）"""
    q = db.query(SalaryRecord)
    if date_from and date_to:
        months = set()
        from_year, from_month = int(date_from[:4]), int(date_from[5:7])
        to_year, to_month = int(date_to[:4]), int(date_to[5:7])
        for y in range(from_year, to_year + 1):
            m_start = from_month if y == from_year else 1
            m_end = to_month if y == to_year else 12
            for m in range(m_start, m_end + 1):
                months.add(f"{y}-{m:02d}")
        if months:
            q = q.filter(or_(*[SalaryRecord.month.like(m + '%') for m in sorted(months)]))
    elif date_from:
        q = q.filter(SalaryRecord.month >= date_from[:7])
    elif date_to:
        q = q.filter(SalaryRecord.month <= date_to[:7])
    elif month:
        q = q.filter(SalaryRecord.month.like(month[:7] + '%'))
    if worker_name:
        w = db.query(SalaryWorker).filter(SalaryWorker.name == worker_name).first()
        if w: q = q.filter(SalaryRecord.worker_id == w.id)
    count = q.filter(SalaryRecord.paid == 0).update({"paid": 1}, synchronize_session=False)
    db.commit()
    return {"ok": True, "updated": count}


@app.get("/api/salary/summary")
def salary_summary(month: str = "", date_from: str = "", date_to: str = "", db: Session = Depends(get_db)):
    """按工人汇总工资，支持月份或日期范围筛选"""
    q = db.query(SalaryRecord)
    if date_from and date_to:
        # month 字段存 YYYY-MM 格式，日期选择器给 YYYY-MM-DD，用月份前缀匹配
        months = set()
        from_year, from_month = int(date_from[:4]), int(date_from[5:7])
        to_year, to_month = int(date_to[:4]), int(date_to[5:7])
        for y in range(from_year, to_year + 1):
            m_start = from_month if y == from_year else 1
            m_end = to_month if y == to_year else 12
            for m in range(m_start, m_end + 1):
                months.add(f"{y}-{m:02d}")
        if months:
            q = q.filter(or_(*[SalaryRecord.month.like(m + '%') for m in sorted(months)]))
    elif date_from:
        q = q.filter(SalaryRecord.month >= date_from[:7])
    elif date_to:
        q = q.filter(SalaryRecord.month <= date_to[:7])
    elif month:
        q = q.filter(SalaryRecord.month.like(month[:7] + '%'))
    rows = q.all()
    workers = {}
    for r in rows:
        wn = r.worker.name if r.worker else "未知"
        if wn not in workers:
            workers[wn] = {"worker_name": wn, "total": 0, "paid_amount": 0, "items": 0}
        workers[wn]["total"] += float(r.amount)
        workers[wn]["items"] += 1
        if r.paid: workers[wn]["paid_amount"] += float(r.amount)
    return sorted(workers.values(), key=lambda x: x["total"], reverse=True)

@app.get("/api/salary/share-image")
def salary_share_image(month: str = "", worker: str = "", db: Session = Depends(get_db)):
    """生成工资单PNG图片，可直接分享到微信"""
    try:
        from PIL import Image, ImageDraw, ImageFont as PILFont
    except ImportError:
        return Response(content=b"Pillow not installed", media_type="text/plain", status_code=500)

    import os as _os2

    q = db.query(SalaryRecord).filter(SalaryRecord.month.like(month[:7] + '%'))
    records = q.all()
    worker_records = [r for r in records if r.worker.name == worker]
    if not worker_records:
        return Response(content=b"no data", media_type="image/png")

    total = sum(float(r.amount) for r in worker_records)
    paid = sum(float(r.amount) for r in worker_records if r.paid)

    # 加载中文字体
    font = None
    font_paths = [
        "/System/Library/Fonts/PingFang.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
        "/usr/share/fonts/truetype/arphic/uming.ttc",
        "/usr/share/fonts/truetype/arphic/ukai.ttc",
    ]
    for fp in font_paths:
        if _os2.path.exists(fp):
            try:
                font = PILFont.truetype(fp, 14)
                font_title = PILFont.truetype(fp, 22)
                font_small = PILFont.truetype(fp, 12)
                break
            except:
                pass

    if not font:
        # 没有中文字体，回退到HTML方式
        return Response(content=f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>工资单</title>
<style>body{{font-family:sans-serif;padding:20px;-webkit-user-select:none;user-select:none}}table{{border-collapse:collapse;width:100%}}
th,td{{border:1px solid #ddd;padding:8px 10px}}th{{background:#f5f5f5}}
.total{{font-size:18px;font-weight:700;margin-top:16px;text-align:right}}</style></head><body>
<h2>{worker} - {month} 工资单</h2><table><tr><th>规格</th><th>颜色</th><th>工序</th><th>数量</th><th>单价</th><th>金额</th><th>支付</th><th>时间</th></tr>
{"".join(f"<tr><td>{r.spec_name or ''}</td><td>{r.color_name or ''}</td><td>{r.item_name}</td><td>{r.quantity}</td><td>¥{float(r.unit_price):.2f}</td><td>¥{float(r.amount):.2f}</td><td>{r.payment_method}</td><td style='font-size:11px'>{r.created_at.strftime('%m-%d %H:%M') if r.created_at else ''}</td></tr>" for r in worker_records)}
</table><div class="total">合计: ¥{total:.2f} | 已付: ¥{paid:.2f} | 未付: ¥{total-paid:.2f}</div>
<p style="text-align:center;color:#999;margin-top:20px">iPhone: 电源+音量+ | 安卓: 电源+音量- — 淼伊库服饰</p></body></html>""",
            media_type="text/html; charset=utf-8")

    try:
        row_h = 30; w = 520; h = 120 + len(worker_records) * row_h + 50
        img = Image.new('RGB', (w, h), 'white')
        draw = ImageDraw.Draw(img)

        draw.text((w/2-80, 16), f"{worker} - {month}工资单", fill='#1e293b', font=font_title)
        draw.text((w/2-60, 48), "淼伊库服饰有限公司", fill='#64748b', font=font_small)

        cols_x = [20, 180, 260, 340, 420]
        headers = ["工序", "数量", "单价", "金额", "支付"]
        y = 80
        draw.rectangle([0, y, w, y+row_h], fill='#f8fafc')
        for hdr, cx in zip(headers, cols_x):
            draw.text((cx, y+7), hdr, fill='#64748b', font=font_small)

        for r in worker_records:
            y += row_h
            vals = [r.item_name, str(r.quantity), f"¥{float(r.unit_price):.2f}", f"¥{float(r.amount):.2f}", r.payment_method]
            for v, cx in zip(vals, cols_x):
                draw.text((cx, y+7), v, fill='#1e293b', font=font)

        y += row_h
        draw.rectangle([0, y, w, y+row_h], fill='#f0fdf4')
        draw.text((20, y+7), f"合计: ¥{total:.2f}  |  已付: ¥{paid:.2f}  |  未付: ¥{total-paid:.2f}", fill='#22c55e', font=font_title)

        buf = io.BytesIO(); img.save(buf, format='PNG'); buf.seek(0)
        return Response(content=buf.getvalue(), media_type="image/png",
                        headers={"Content-Disposition": f'inline; filename="salary_{worker}_{month}.png"'})
    except Exception as e:
        return Response(content=f"生成失败: {e}".encode(), media_type="text/plain", status_code=500)


@app.get("/api/salary/share")
def salary_share(month: str = "", worker: str = "", db: Session = Depends(get_db)):
    """工人工资单（无需登录）
    - 指定 month: 查看该月明细（管理员分享）
    - 不指定 month: 查看所有未付款记录 + 二维码（工人自行扫码）
    """
    from urllib.parse import quote

    # 永久二维码 URL（不带 month，工人可保存）
    perm_url = f"http://47.96.91.217/api/salary/share?worker={quote(worker)}"

    if month:
        # 管理员查看指定月份
        q = db.query(SalaryRecord).filter(SalaryRecord.month.like(month[:7] + '%'))
        records = q.all()
        worker_records = [r for r in records if r.worker.name == worker]
        page_title = f"{worker} - {month}工资单"
        page_sub = f"{month} &nbsp;|&nbsp; 淼伊库服饰有限公司"
        is_worker_view = False
    else:
        # 工人扫码查看：只显示未付款
        all_records = db.query(SalaryRecord).filter(SalaryRecord.paid == False).all()
        worker_records = [r for r in all_records if r.worker.name == worker]
        worker_records.sort(key=lambda r: r.created_at or datetime.min, reverse=True)
        page_title = f"{worker} - 工资查询"
        page_sub = "未付款记录 &nbsp;|&nbsp; 淼伊库服饰有限公司"
        is_worker_view = True

    # 工人扫码查看时记录活跃度
    if is_worker_view and worker:
        try:
            db.add(SalaryViewLog(worker_name=worker, month=month or ""))
            db.commit()
        except Exception:
            db.rollback()

    if not worker_records:
        return Response(content=f'<html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head><body style="font-family:PingFang SC,Microsoft YaHei,sans-serif;max-width:600px;margin:0 auto;padding:20px;text-align:center"><h2>💰 {worker}</h2><p style="color:#22c55e;font-size:18px">🎉 没有未付款记录</p><p style="color:#94a3b8">所有工资已结清</p><p style="color:#64748b">淼伊库服饰有限公司</p></body></html>', media_type="text/html")

    total = sum(float(r.amount) for r in worker_records)
    paid = sum(float(r.amount) for r in worker_records if r.paid)

    rows_html = ""
    for r in worker_records:
        pay_color = "#22c55e" if r.paid else "#ef4444"
        pay_text = "已付" if r.paid else "未付"
        r_time = r.created_at.strftime('%m-%d %H:%M') if r.created_at else ''
        r_month = r.month if r.month else ''
        rows_html += f"<tr><td>{r.spec_name or ''}</td><td>{r.color_name or ''}</td><td>{r.item_name}</td><td>{r.quantity}</td><td>¥{float(r.unit_price):.2f}</td><td>¥{float(r.amount):.2f}</td><td style='color:{pay_color};font-weight:600'>{pay_text}</td><td style='font-size:11px'>{r_time}</td><td style='font-size:11px;color:#94a3b8'>{r_month}</td></tr>"

    qr_img_html = ""
    try:
        import qrcode as qrcode_lib
        import base64
        img = qrcode_lib.make(perm_url)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        qr_b64 = base64.b64encode(buf.getvalue()).decode()
        qr_img_html = f'<div class="qr-box"><p style="font-weight:600;margin-bottom:8px">📱 扫码随时查看工资</p><img src="data:image/png;base64,{qr_b64}" style="width:180px;height:180px" alt="QR码"><p style="font-size:11px;color:#94a3b8;margin-top:6px">长按保存二维码，随时扫码查看</p></div>'
    except Exception as e:
        print(f"[QR] failed: {e}")
        qr_img_html = f'<div class="qr-box"><p style="color:#ef4444">⚠️ 二维码生成失败</p></div>'

    html = f"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{page_title}</title>
<style>body{{font-family:'PingFang SC','Microsoft YaHei',sans-serif;max-width:600px;margin:0 auto;padding:16px;color:#1e293b;-webkit-user-select:none;user-select:none;background:#fff}}
.header{{text-align:center;padding:20px 0;border-bottom:2px solid #4f6ef7;margin-bottom:16px}}
.header h2{{margin:0;font-size:20px}}.header .sub{{color:#64748b;font-size:13px;margin-top:4px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#f8fafc;color:#64748b;font-weight:600;padding:10px 8px;border-bottom:2px solid #e2e8f0;font-size:12px}}
td{{padding:10px 8px;border-bottom:1px solid #f1f5f9}}
tr:nth-child(even) td{{background:#fafbfd}}
.total{{text-align:right;font-size:16px;font-weight:700;padding:16px 0;border-top:2px solid #4f6ef7;margin-top:8px}}
.total span{{color:#4f6ef7}}
.qr-box{{text-align:center;padding:20px;margin-top:20px;border-top:2px dashed #e2e8f0}}
.tip{{background:#f8fafc;border-radius:10px;padding:12px;text-align:center;font-size:13px;color:#64748b;margin-top:12px}}
.footer{{text-align:center;color:#94a3b8;font-size:11px;margin-top:12px;padding-top:12px;border-top:1px solid #f1f5f9}}
@media print{{body{{margin:0;padding:10px}}.qr-box{{page-break-before:always}}}}
</style></head><body>
<div class="header"><h2>💰 {worker} 工资单</h2><div class="sub">{page_sub}</div></div>
<table><thead><tr><th>规格</th><th>颜色</th><th>工序</th><th>数量</th><th>单价</th><th>金额</th><th>支付</th><th>时间</th><th>月份</th></tr></thead><tbody>{rows_html}</tbody></table>
<div class="total">合计 <span>¥{total:.2f}</span> &nbsp; 已付 <span>¥{paid:.2f}</span> &nbsp; 未付 <span style="color:#ef4444">¥{total-paid:.2f}</span></div>
{qr_img_html}
<div class="tip">📱 保存二维码随时查看 &nbsp;|&nbsp; 自动显示未付款记录 &nbsp;|&nbsp; 淼伊库服饰有限公司</div>
<div class="footer">生成时间: {date.today()}</div>
</body></html>"""
    return Response(content=html, media_type="text/html; charset=utf-8")


@app.get("/api/salary/views")
def salary_views(db: Session = Depends(get_db), admin: Admin = Depends(get_current_admin)):
    """工资查看活跃度统计"""
    # 所有活跃工人
    workers = db.query(SalaryWorker).filter(SalaryWorker.is_active == 1).all()

    # 每个工人的查看记录
    result = []
    for w in workers:
        views = db.query(SalaryViewLog).filter(SalaryViewLog.worker_name == w.name).all()
        last_view = max((v.viewed_at for v in views), default=None)
        result.append({
            "worker_name": w.name,
            "job_type": w.job_type or "",
            "view_count": len(views),
            "last_view": last_view.strftime("%Y-%m-%d %H:%M") if last_view else None,
        })

    # 排序：没看过的排前面，然后按最后查看时间从旧到新
    result.sort(key=lambda x: (x["last_view"] is None, x["last_view"] or ""))

    total = len(result)
    viewed = sum(1 for r in result if r["view_count"] > 0)
    not_viewed = total - viewed

    return {
        "items": result,
        "summary": {
            "total_workers": total,
            "viewed": viewed,
            "not_viewed": not_viewed,
            "view_rate": round(viewed / total * 100, 1) if total else 0,
        }
    }


@app.get("/api/salary/export")
def salary_export(month: str = "", db: Session = Depends(get_db)):
    """导出工资表为Excel"""
    q = db.query(SalaryRecord).order_by(SalaryRecord.worker_id, SalaryRecord.id)
    if month: q = q.filter(SalaryRecord.month.like(month[:7] + '%'))
    rows = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "工资明细"
    ws.append(["工人", "月份", "工序", "数量", "单价", "金额", "支付方式", "已付", "备注"])
    for r in rows:
        ws.append([r.worker.name if r.worker else "", r.month, r.item_name,
                   r.quantity, float(r.unit_price), float(r.amount),
                   r.payment_method, "是" if r.paid else "否", r.remark])
    # 汇总
    summary = salary_summary(month, db)
    ws.append([])
    ws.append(["工人", "", "", "", "", "合计", "", "", ""])
    for s in summary:
        ws.append([s["worker_name"], "", "", "", "", s["total"], "", "", ""])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=salary_{month}.xlsx"})


# ============================================================
# 首页仪表盘数据
# ============================================================
@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db)):
    today = date.today()
    default_threshold = get_low_stock_threshold(db)

    # 在职销售人员ID集合
    active_sp_ids = set(s.id for s in db.query(Salesperson.id).filter(Salesperson.is_active == 1).all())

    today_rows = db.query(Transaction).filter(Transaction.trans_date == today).all()
    today_out = sum(float(t.amount) for t in today_rows if t.trans_type == "出库" and (not t.salesperson_id or t.salesperson_id in active_sp_ids))
    today_out_qty = sum(t.quantity for t in today_rows if t.trans_type == "出库")
    today_in = sum(t.quantity for t in today_rows if t.trans_type == "入库")

    # 本月累计（只统计在职销售人员）
    month_start = today.replace(day=1)
    month_rows = db.query(Transaction).filter(
        Transaction.trans_date >= month_start,
        Transaction.trans_date <= today,
        Transaction.trans_type == "出库",
    ).all()
    month_total = sum(float(t.amount) for t in month_rows if not t.salesperson_id or t.salesperson_id in active_sp_ids)
    month_out_qty = sum(t.quantity for t in month_rows)

    # 本月累计入库
    month_in = db.query(func.sum(Transaction.quantity)).filter(
        Transaction.trans_date >= month_start,
        Transaction.trans_date <= today,
        Transaction.trans_type == "入库",
    ).scalar() or 0

    # 低库存
    low_skus = db.query(ProductSku).filter(
        ProductSku.current_stock < func.coalesce(ProductSku.low_stock_threshold, default_threshold)
    ).count()

    # 当月工资汇总
    salary_month = db.query(func.sum(SalaryRecord.amount)).filter(
        SalaryRecord.month.like(today.strftime("%Y-%m") + '%')
    ).scalar() or 0

    return {
        "today_out_amount": today_out,
        "today_out_qty": today_out_qty,
        "today_in_qty": today_in,
        "month_total": month_total,
        "month_out_qty": month_out_qty,
        "month_in_qty": int(month_in),
        "salary_month_total": float(salary_month),
        "low_stock_count": low_skus,
        "total_specs": db.query(ProductSpec).filter(ProductSpec.is_active == 1).count(),
        "total_colors": db.query(Color).filter(Color.is_active == 1).count(),
        "total_workers": db.query(SalaryWorker).filter(SalaryWorker.is_active == 1).count(),
        "total_price_items": db.query(SalaryPrice).filter(SalaryPrice.is_active == 1).count(),
    }


# ============================================================
# 静态文件 & 前端页面
# ============================================================
app.mount("/static", StaticFiles(directory="static", html=True), name="static")


@app.get("/")
def index():
    return FileResponse("static/index.html")


# ============================================================
# 启动入口
# ============================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
